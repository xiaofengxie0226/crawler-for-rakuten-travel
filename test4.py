import requests, bs4, openpyxl, gevent, time, socket
from openpyxl import Workbook
from gevent import monkey
from gevent.queue import Queue

work = Queue()
monkey.patch_all()
headers = {
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/78.0.3904.108 Safari/537.36'}

socket.setdefaulttimeout(20)
wb = openpyxl.load_workbook(filename=r'C:\Users\JFtra\Desktop\hotel-all.xlsx')
sheetID = wb.get_sheet_by_name('all')

ID = []
hotel_ID = {}
after = []
tasks_list = []


for a in range(2, 4):  # 获取每个酒店品牌的ID
    ID.append(str(sheetID.cell(row=a, column=1).value))


def hotel_crawler():
    while not work.empty():
        try:
            try:
                try:
                    start = time.time()
                    url = work.get_nowait()
                    r = requests.get(url, headers=headers)
                    # print(url, work.qsize(), r.status_code)
                    ab = bs4.BeautifulSoup(r.text, 'html.parser')
                    try:
                        ac = ab.find('h2', class_="commentTitle")
                        next_url = ac.find('a')['href']
                        r2 = requests.get(next_url, headers=headers)
                        print(next_url, work.qsize(), r2.status_code)
                        try:
                            sheet_hotel['A1'] = '宿名'
                            sheet_hotel['B1'] = '名前'
                            sheet_hotel['C1'] = 'コメント'
                            sheet_hotel['D1'] = '総合'
                            sheet_hotel['E1'] = 'サービス'
                            sheet_hotel['F1'] = '立地'
                            sheet_hotel['G1'] = '部屋'
                            sheet_hotel['H1'] = '設備'
                            sheet_hotel['I1'] = '風呂'
                            sheet_hotel['J1'] = '食事'
                            sheet_hotel['K1'] = '宿泊年月'
                            sheet_hotel['L1'] = '目的'
                            sheet_hotel['M1'] = '同伴者'
                            sheet_hotel['N1'] = 'コメント時間'
                            sheet_hotel['O1'] = 'ID'
                            n = 0
                            t = 0
                            while True:
                                n = n + 1
                                t = t + 1
                                print('comment:{}'.format(n))
                                if t > 2000:
                                    print('it needs a save')
                                    excel.save(
                                        filename=r'C:\Users\JFtra\.PyCharmCE2018.3\config\scratches\hotel-t1.xlsx')
                                    time.sleep(3)
                                    t = 0
                                try:
                                    res = requests.get(next_url, headers=headers, stream=True)
                                    bb = bs4.BeautifulSoup(res.text, 'html.parser')
                                    point = []
                                    status = []
                                    bs = bb.find('div', id="RthNameArea")
                                    hotel = bs.find('a', class_="rtconds fn").text
                                    bc = bb.find('dl', class_="commentReputation")
                                    name = bc.find('span', class_="user").text
                                    time_comment = bc.find('span', class_="time").text
                                    comment = bc.find('p', class_="commentSentence").text
                                    bd = bb.find('dl', class_="commentPurpose")
                                    h_id = hotel_ID[hotel]
                                    try:
                                        total = bd.find('span', class_="rate rate50").text
                                    except AttributeError:
                                        total = '-'
                                    try:
                                        for mark in bd.find_all('li'):
                                            s = mark.find('span', class_="rate").text
                                            point.append([s])
                                        str = ''
                                        service = str.join(point[1])
                                        place = str.join(point[2])
                                        room = str.join(point[3])
                                        device = str.join(point[4])
                                        bath = str.join(point[5])
                                        supper = str.join(point[6])
                                        for s in bd.find_all('dd'):
                                            suit = s.text
                                            status.append([suit])
                                        str = ''
                                        try:
                                            check = str.join(status[-1])
                                        except IndexError:
                                            check = '/'
                                        try:
                                            purpose = str.join(status[1])
                                            mate = str.join(status[2])
                                        except IndexError:
                                            purpose = '/'
                                            mate = '/'
                                        try:
                                            sheet_hotel.append(
                                                [hotel, name, comment, total, service, place, room, device, bath,
                                                 supper,
                                                 purpose, mate, check,
                                                 time_comment, h_id])
                                        except openpyxl.utils.exceptions.IllegalCharacterError:
                                            print('IllegalCharacterError')
                                        try:
                                            bs = bb.find('li', class_="linkNext")
                                            next_url = bs.find('a')['href']
                                        except AttributeError:
                                            # print('1,No next url.ID:{}, comment:{}'.format(hotel,time_comment))
                                            print('2,No next url.ID:{}, comment:{}'.format(hotel, time_comment))
                                            break
                                    except AttributeError:
                                        print('Missing data')
                                        service = place = room = device = bath = supper = '-'
                                        purpose = mate = '/'
                                        for s in bb.find_all('dd'):
                                            suit = s.text
                                            status.append([suit])
                                        str = ''
                                        try:
                                            check = str.join(point[-1])
                                            sheet_hotel.append(
                                                [hotel, name, comment, total, service, place, room, device, bath,
                                                 supper, check,
                                                 purpose,
                                                 mate,
                                                 time_comment, h_id])
                                        except IndexError:
                                            check = '/'
                                            sheet_hotel.append(
                                                [hotel, name, comment, total, service, place, room, device, bath,
                                                 supper, check,
                                                 purpose,
                                                 mate,
                                                 time_comment, h_id])
                                    try:
                                        bs = bb.find('li', class_="linkNext")
                                        next_url = bs.find('a')['href']
                                    except AttributeError:
                                        print('finish2.ID:{}', format(hotel))
                                        break
                                except requests.exceptions.ConnectionError:
                                    requests.status_code = "Connection refused"
                                    continue
                            m = 0
                            m = m + n
                            end = time.time()
                            print(end - start)
                            print('previous comment:{}'.format(m))
                            excel.save(
                                filename=r'C:\Users\JFtra\.PyCharmCE2018.3\config\scratches\hotel-{}.xlsx'.format(
                                    ID[i]))
                            excel.close()
                        except:
                            continue
                    except AttributeError:
                        continue
                except requests.exceptions.ConnectionError:
                    time.sleep(5)
            except requests.exceptions.InvalidURL:
                print('invalid')
        except requests.exceptions.MissingSchema:
            print('None')


for i in range(len(ID)):  # 获取每个要爬取的网站
    sheetI = wb.get_sheet_by_name(ID[i])
    print(ID[i])
    long = sheetI.max_row
    for j in range(2, long + 1):
        url = sheetI.cell(row=j, column=4).value
        hotel_ID[sheetI.cell(row=j, column=2).value] = sheetI.cell(row=j, column=1).value
        # print(type(url))
        if url == '':
            print('no url')
        else:
            work.put_nowait(url)
    print(work.qsize(),long)
    excel = Workbook()
    sheet_hotel = excel.active
    for x in range(long):
        task = gevent.spawn(hotel_crawler)
        tasks_list.append(task)
        work.put_nowait(hotel_crawler)
    gevent.joinall(tasks_list)
    excel.save(filename=r'C:\Users\JFtra\.PyCharmCE2018.3\config\scratches\hotel-{}.xlsx'.format(ID[i]))
    excel.close()
    print('{} is over'.format(ID[i]))
    hotel_ID.clear()
print(after)
localtime = time.localtime(time.time())
print(localtime)
